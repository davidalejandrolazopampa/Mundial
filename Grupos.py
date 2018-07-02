#importamos el tkinter(ventana) y panda(excel)
from tkinter import *
from tkinter import ttk
from pandas import *
from openpyxl import load_workbook
window=Tk()
#funciones de btns
def buscar():
    pestasec=Tk()
    #pestasec.geometry("160x170+0+0")
    pestasec.configure(background='lime')
    x=buscar_text.get()
    Datos=xls.loc[x]
    l10=Label(pestasec,text="Busqueda de Pais")
    l10.grid(row=0,column=0)
    l10=Label(pestasec,text=Datos,bg="lime")
    l10.grid(row=1,column=0)
def buscar1():
    pestasec=Tk()
    pestasec.geometry("220x130+0+0")
    pestasec.configure(background='lime')
    x=buscar_textG.get()
    Datos=xls1.loc[x]
    l10=Label(pestasec,text="Busqueda de Grupos")
    l10.grid(row=0,column=0)
    l10=Label(pestasec,text=Datos,bg="lime")
    l10.grid(row=1,column=0)
def mostrar():
    pestasec=Tk()
    xls=ExcelFile("mundialgrupos.xlsx").parse('Hoja1')
    xls.set_index("Paises",inplace=True)
    l11=Label(pestasec,text="Paises del mundial- Rusia 2018")
    l11.grid(row=0,column=0)
    l10=Label(pestasec,text=xls,bg="lightskyblue",anchor="w")
    l10.grid(row=1,column=0)
    pestasec.mainloop()
def agregar():
    doc="mundialgrupos.xlsx"
    wb=load_workbook(doc)
    sheet=wb.active
    u=34

    q=pais_text.get()
    l=grupo_text.get()
    e=pj_text.get()
    w=g_text.get()
    k=e_text.get()
    f=p_text.get()
    z=dg_text.get()
    p=pts_text.get()


    #sheet.cell(row=31, column=1).value=u
    sheet.cell(row=u, column=1).value=q
    sheet.cell(row=u, column=2).value=l
    sheet.cell(row=u, column=3).value=e
    sheet.cell(row=u, column=4).value=w
    sheet.cell(row=u, column=5).value=k
    sheet.cell(row=u, column=6).value=f
    sheet.cell(row=u, column=7).value=z
    sheet.cell(row=u, column=8).value=p
    wb.save(doc)

window.maxsize(width=715, height=430)
window.minsize(width=715, height=430)
window.title("Grupos del Mundial 2018")

Imagen=PhotoImage(file="img/111.gif")
fondo=Label(window,image=Imagen).place(x=0,y=0)

window.title("Grupos del Mundial 2018")

xls=ExcelFile("mundialgrupos.xlsx").parse('Hoja1')
xls.set_index("Paises",inplace=True)

xls1=ExcelFile("mundialgrupos.xlsx").parse('Hoja1')
xls1.set_index("Grupos",inplace=True)

#caja de texto

buscar_text=StringVar()
e1=Entry(window,textvariable=buscar_text,width=24)
e1.place(x=275,y=285)


buscar_textG=ttk.Combobox(window,width=24)
buscar_textG.place(x=275,y=315)
buscar_textG.set('A')
buscar_textG['values']=('A','B','C','D','E','F','H')
#botones locos

b1=Button(window,text='Buscar Pais',width=15,command=buscar)
b1.place(x=450,y=285)

b1=Button(window,text='Buscar Grupo',width=15,command=buscar1)
b1.place(x=450,y=315)

b1=Button(window,text='Mostrar Total',width=15,command=mostrar)
b1.place(x=150,y=285)

b1=Button(window,text='SALIR',width=15,command=quit,bg='white')
b1.place(x=610,y=405)
#------------------------------------------------------------------------
l1=Label(window,text="Numero")
l1.grid(row=0,column=0)

l1=Label(window,text="Paises")
l1.grid(row=0,column=1)

l1=Label(window,text="Grupo")
l1.grid(row=0,column=2)

l1=Label(window,text="PJ")
l1.grid(row=0,column=3)

l1=Label(window,text="G")
l1.grid(row=0,column=4)

l1=Label(window,text="E")
l1.grid(row=0,column=5)

l1=Label(window,text="P")
l1.grid(row=0,column=6)

l1=Label(window,text="DG")
l1.grid(row=0,column=7)

l1=Label(window,text="Pts")
l1.grid(row=0,column=8)

cantidad=int(len(xls)+1)
numero_pais=StringVar()
numero_pais.set(cantidad)
e1=Entry(window,textvariable=numero_pais,width=2)
e1.grid(row=1,column=0)

pais_text=StringVar()
e1=Entry(window,textvariable=pais_text,width=10)
e1.grid(row=1,column=1)

grupo_text=ttk.Combobox(window,width=10)
grupo_text.grid(row=1,column=2)
grupo_text.set('A')
grupo_text['values']=('A','B','C','D','E','F','H')

pj_text=StringVar()
e1=Entry(window,textvariable=pj_text,width=2)
e1.grid(row=1,column=3)

g_text=StringVar()
e1=Entry(window,textvariable=g_text,width=2)
e1.grid(row=1,column=4)

e_text=StringVar()
e1=Entry(window,textvariable=e_text,width=2)
e1.grid(row=1,column=5)

p_text=StringVar()
e1=Entry(window,textvariable=p_text,width=2)
e1.grid(row=1,column=6)

dg_text=StringVar()
e1=Entry(window,textvariable=dg_text,width=2)
e1.grid(row=1,column=7)

pts_text=StringVar()
e1=Entry(window,textvariable=pts_text,width=2)
e1.grid(row=1,column=8)


b1=Button(window,text='Agregar',width=12,command=agregar)
b1.grid(row=1,column=10)

b1=Button(window,text='Editar',width=12)
b1.grid(row=1,column=11)

b1=Button(window,text='Eliminar',width=12)
b1.grid(row=1,column=12)
window.mainloop()

